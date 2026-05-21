#!/usr/bin/env python3
"""Excel (UpperLowerBody hypertrophy 9weeks) -> JSON converter.

Usage:
    python3 scripts/extract_data.py [path/to/source.xlsx]

Default source: ../[Ver2.2]UpperLowerBodyhypertrophy 9weeks [4].xlsx
Outputs: src/data/{program,personal,summary}.json
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime, time
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT.parent / "[Ver2.2]UpperLowerBodyhypertrophy 9weeks [4].xlsx"
OUT_DIR = ROOT / "src" / "data"


def interval_to_seconds(v: Any) -> int | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        t = v.time()
        return t.hour * 3600 + t.minute * 60 + t.second
    if isinstance(v, time):
        return v.hour * 3600 + v.minute * 60 + v.second
    if isinstance(v, (int, float)):
        return int(float(v) * 86400)
    return None


def num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_personal(ws) -> dict:
    def g(coord):
        return ws[coord].value

    return {
        "big3": {
            "squat": num(g("B9")),
            "bench": num(g("C9")),
            "deadlift": num(g("D9")),
            "highbarSquat": num(g("E9")),
            "shoulderPress": num(g("F9")),
        },
        "profile": {
            "heightCm": num(g("C14")),
            "weightKg": num(g("C15")),
            "mode": g("C16"),
            "dailyIntensity": g("C17"),
            "trainingLevel": g("C18"),
        },
        "nutrition": {
            "maintenanceKcal": num(g("C19")),
            "overKcal": num(g("C20")),
            "underKcal": num(g("C21")),
            "proteinPerMeal": num(g("C22")),
            "targetKcal": num(g("C24")),
            "p": num(g("C25")),
            "f": num(g("C26")),
            "c": num(g("C27")),
        },
    }


def parse_summary(ws) -> list[dict]:
    weeks: dict[int, dict] = {i: {"week": i} for i in range(1, 10)}
    for row in ws.iter_rows(values_only=False):
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            m = re.fullmatch(r"WEEK(\d+)", v)
            if not m:
                continue
            wk = int(m.group(1))
            d_val = num(ws.cell(row=c.row, column=4).value)
            if d_val is None:
                continue
            for r in range(c.row - 1, 0, -1):
                hv = ws.cell(row=r, column=3).value
                if not isinstance(hv, str):
                    continue
                if hv.startswith("TOTAL VOLUME(Set"):
                    weeks[wk]["totalSets"] = d_val
                    break
                if hv.startswith("INTENSITY AVARAGE"):
                    weeks[wk]["avgRpe"] = d_val
                    break
                if hv.startswith("TOTAL VOLUME(REPS"):
                    weeks[wk]["totalReps"] = d_val
                    break
                if hv.startswith("TOTAL VOLUME(KG"):
                    weeks[wk]["totalKg"] = d_val
                    break
    return [weeks[i] for i in range(1, 10)]


def parse_week(ws, week_no: int) -> dict:
    days: list[dict] = []
    current_day: dict | None = None
    current_exercise: dict | None = None

    for row in ws.iter_rows(values_only=False):
        a = row[0].value
        b = row[1].value if len(row) > 1 else None
        c = row[2].value if len(row) > 2 else None

        if isinstance(a, str) and a.strip().startswith("Day "):
            current_day = {
                "id": a.strip(),
                "label": a.strip(),
                "type": "Upper" if "Upper" in a else ("Lower" if "Lower" in a else "Other"),
                "exercises": [],
            }
            days.append(current_day)
            current_exercise = None
            continue

        if current_day is None:
            continue

        if isinstance(b, str) and b.strip() == "種目":
            continue
        if isinstance(c, str) and c.strip() == "TOTAL":
            current_exercise = None
            continue

        if isinstance(b, str) and b.strip():
            name = b.strip()
            rpe = num(row[2].value)
            pct = num(row[3].value)
            sets = num(row[4].value)
            rep_range = row[5].value
            if isinstance(rep_range, (int, float)):
                rep_range = str(int(rep_range))
            interval = interval_to_seconds(row[6].value)
            ref_kg = num(row[7].value)
            note = row[16].value if len(row) > 16 else None
            if isinstance(note, (int, float)) and note == 0:
                note = None

            if name == "Backoff" and current_exercise is not None:
                current_exercise["backoff"] = {
                    "rpe": rpe,
                    "percentRm": pct,
                    "sets": sets,
                    "repRange": rep_range,
                    "intervalSec": interval,
                    "refKg": ref_kg,
                }
                continue

            current_exercise = {
                "part": (a or "").strip() if isinstance(a, str) else None,
                "name": name,
                "rpe": rpe,
                "percentRm": pct,
                "sets": sets,
                "repRange": rep_range,
                "intervalSec": interval,
                "refKg": ref_kg,
                "note": note if isinstance(note, str) else None,
                "backoff": None,
            }
            current_day["exercises"].append(current_exercise)

    return {"week": week_no, "days": days}


def scrub(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {k: scrub(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [scrub(v) for v in obj]
    if isinstance(obj, datetime):
        return obj.isoformat()
    if isinstance(obj, time):
        return obj.isoformat()
    return obj


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"ERROR: not found: {xlsx_path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    def dump(obj, name):
        (OUT_DIR / name).write_text(
            json.dumps(scrub(obj), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    dump(parse_personal(wb["PersonalInfo"]), "personal.json")
    dump(parse_summary(wb["TotalVolume"]), "summary.json")

    weeks_out = []
    for i in range(1, 10):
        sn = f"WEEK{i:02d}"
        weeks_out.append(parse_week(wb[sn], i))
    dump(weeks_out, "program.json")

    total_ex = sum(len(d["exercises"]) for w in weeks_out for d in w["days"])
    print(f"wrote {OUT_DIR}/personal.json, summary.json, program.json")
    print(f"weeks={len(weeks_out)} total_exercises={total_ex}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
